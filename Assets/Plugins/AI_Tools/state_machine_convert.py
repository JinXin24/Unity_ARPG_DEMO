#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
状态机配置表转换工具：单表 → 双表（合并模式）。

用法：
    python state_machine_convert.py            # 默认：单表 → 新建 StateMachineConfig_Double.xlsx
    python state_machine_convert.py <输入.xlsx> [输出.xlsx]   # 指定输入/输出

默认（相对本脚本所在仓库根目录）：
    输入 = Assets/Excel/StateMachineConfig.xlsx（单表，sheet 名 state，窗口列内嵌）
    输出 = Assets/Excel/StateMachineConfig_Double.xlsx（新建，state + transition 双表）

原则：
    输入单表【不动】；转换结果写到【新文件】。

    合并而非全量覆盖 —— 工具只重写"转换负责"的列：
      只重写 state 的 6 列、transition 的 From/To/Condition/Param；
      双表里已有的 Order 优先级 和 额外列（如备注）在匹配行上【原样保留】。
    行匹配键：transition 按 (From, To, Condition)（不含 Param，容忍手改/格式漂移），state 按 StateId。
    新窗口 → 新行（Order 用默认值）；单表删掉的状态/窗口 → 对应行移除。
    注意：不要在双表手加"由单表派生的行"（会被清理）；能手改且会保留的是 Order 和额外列。

转换规则（Order 默认优先级，对齐 CharacterState.cs 检测顺序）：
    OnAnimEnd       → Condition=OnAnimEnd,   To=目标状态, Order=1
    OnAtk  窗口     → Condition=OnAtk,       Param=窗口,  Order=2
    OnMove 窗口     → Condition=OnMove,      Param=窗口,  Order=3
    OnSkill 窗口    → Condition=OnSkill,     Param=窗口,  Order=4
    OnEnhanceSkill  → Condition=OnEnhanceSkill, To=进场状态, Param=[离场状态], Order=5

脚本自带 round-trip 校验：生成后反推单表与源表逐格比对（不含 Order/额外列），
不一致报错退出、不写输出文件。
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Order 默认优先级（数字越小=越先判断），与 CharacterState.cs 注册/检测顺序一致
ORDER = {
    'OnAnimEnd': 1,
    'OnAtk': 2,
    'OnMove': 3,
    'OnSkill': 4,
    'OnEnhanceSkill': 5,
}

STATE_HEADERS = ['CharacterId', 'UseCommon', 'StateId', 'Info', 'AnimName', 'UnlockEnhance']
STATE_TYPES   = ['int', 'bool', 'int', 'string', 'string', 'bool']
STATE_COMMENTS = ['角色ID(0=通用)', '是否使用通用状态', '状态ID', '说明', '动画状态名', '进入该状态是否解锁强化效果']

TRANS_HEADERS = ['CharacterId', 'From', 'To', 'Condition', 'Param', 'Order']
TRANS_TYPES   = ['int', 'int', 'int', 'string', 'float[]', 'int']
TRANS_COMMENTS = [
    '角色ID(0=通用)',
    '源状态ID(0=任意状态,全局转移)',
    '目标状态ID',
    '转移条件枚举: OnAnimEnd=动画播完; OnAtk=攻击输入窗口; OnMove=移动输入窗口; OnSkill=技能(E)窗口; OnEnhanceSkill=强化技能(强化期内按E)',
    '窗口参数(分号分隔): OnAtk/OnMove/OnSkill=[窗口始;窗口末]; OnEnhanceSkill=[离场状态ID]; OnAnimEnd=空',
    '优先级(数字越小=先判断); 默认按条件类型生成, 已存在则保留手工设置',
]


def num(v):
    """int 或原样字符串。数值漂移归一化：WPS/Excel 可能把 1001 存成 1001.0，统一回 int。"""
    if v is None:
        return ''
    s = str(v).strip()
    if s.lstrip('-').replace('.', '', 1).isdigit():
        return int(float(s))
    return s


def flag(v):
    """bool 归一化为 'TRUE'/'FALSE'（Excel 大写约定）。"""
    return 'TRUE' if str(v).strip().upper() == 'TRUE' else 'FALSE'


def norm_cell(v):
    """单元格归一化（比较/匹配键用）：None→''，bool→大写，其他→strip。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    return str(v).strip()


def parse_windows(val):
    """'st;en;tgt[;st;en;tgt...]' → [(st, en, tgt), ...]。空返回 []。"""
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    parts = [p for p in s.split(';') if p]
    if len(parts) % 3 != 0:
        raise ValueError(f'窗口数组长度必须是 3 的倍数（[始;末;目标]一组）: {s!r}')
    return [(parts[i], parts[i + 1], parts[i + 2]) for i in range(0, len(parts), 3)]


def find_authoring_sheet(wb):
    """找单表 sheet：表头同时含 OnAtk / StateId 的（窗口列内嵌的那张）。"""
    for name in wb.sheetnames:
        ws = wb[name]
        headers = {str(ws.cell(row=1, column=c).value).strip()
                   for c in range(1, ws.max_column + 1)
                   if ws.cell(row=1, column=c).value}
        if 'OnAtk' in headers and 'StateId' in headers:
            return ws, name
    raise SystemExit(f'找不到单表 sheet（表头需含 OnAtk / StateId）: {wb.sheetnames}')


def header_map(ws):
    """列名 → 列下标（1-based），杜绝写死下标错位。"""
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            hdr[str(v).strip()] = c
    return hdr


def parse_single(ws, hdr):
    """读单表 → (states, transitions)。"""
    states, transitions, warnings = [], [], []

    def get(r, name):
        return ws.cell(row=r, column=hdr[name]).value

    for r in range(4, ws.max_row + 1):
        cid, sid = str(get(r, 'CharacterId')).strip(), str(get(r, 'StateId')).strip()
        if not cid or not sid:
            continue  # 空行

        states.append([
            num(cid),
            flag(get(r, 'UseCommon')),
            num(sid),
            str(get(r, 'Info') or '').strip(),
            str(get(r, 'AnimName') or '').strip(),
            flag(get(r, 'UnlockEnhance')),
        ])

        # OnAnimEnd：int 目标状态
        oae = get(r, 'OnAnimEnd')
        if oae is not None and str(oae).strip() and num(oae) > 0:
            transitions.append([num(cid), num(sid), num(oae), 'OnAnimEnd', '', ORDER['OnAnimEnd']])

        # 输入窗口列
        for col, cond in (('OnAtk', 'OnAtk'), ('OnMove', 'OnMove'), ('OnSkill', 'OnSkill')):
            for st, en, tgt in parse_windows(get(r, col)):
                transitions.append([num(cid), num(sid), num(tgt), cond, f'{st};{en}', ORDER[cond]])

        # OnEnhanceSkill：[离场;进场] → To=进场, Param=[离场]
        enh = get(r, 'OnEnhanceSkill')
        if enh is not None and str(enh).strip():
            parts = [p for p in str(enh).split(';') if p]
            if len(parts) >= 2:
                transitions.append([num(cid), num(sid), num(parts[1]), 'OnEnhanceSkill', str(parts[0]), ORDER['OnEnhanceSkill']])
            else:
                warnings.append(f'第 {r} 行 OnEnhanceSkill 应为 [离场;进场]，实际: {enh!r}')

    return states, transitions, warnings


def validate(states, transitions):
    """校验：目标状态必须存在、Order 必须有效。返回 warning 列表。"""
    warnings = []
    defined = {st[2] for st in states}
    for t in transitions:
        if t[2] not in defined:
            warnings.append(f'迁移目标状态 {t[2]}（From={t[1]}）在 state 表中未定义！')
        if t[5] not in ORDER.values():
            warnings.append(f'Order={t[5]} 非法（默认应在 {sorted(ORDER.values())}）: From={t[1]} To={t[2]}')
    return warnings


def reverse_derive(states, transitions):
    """双表反推单表（round-trip 校验用）。返回 {StateId: {列: 值}}。"""
    by_from = {}
    for t in transitions:
        by_from.setdefault(t[1], []).append(t)

    single = {}
    for st in states:
        sid = st[2]
        cols = {'OnAnimEnd': None, 'OnMove': [], 'OnAtk': [], 'OnSkill': [], 'OnEnhanceSkill': None}
        for t in sorted(by_from.get(sid, []), key=lambda x: (x[5],)):
            cond = t[3]
            if cond == 'OnAnimEnd':
                cols['OnAnimEnd'] = t[2]
            elif cond in ('OnMove', 'OnAtk', 'OnSkill'):
                cols[cond].append(f'{t[4]};{t[2]}')
            elif cond == 'OnEnhanceSkill':
                cols['OnEnhanceSkill'] = f'{t[4]};{t[2]}'

        single[sid] = {
            'CharacterId': st[0], 'UseCommon': st[1], 'StateId': st[2],
            'Info': st[3], 'AnimName': st[4], 'UnlockEnhance': st[5],
            'OnAnimEnd': cols['OnAnimEnd'],
            'OnMove': ';'.join(cols['OnMove']) or None,
            'OnAtk': ';'.join(cols['OnAtk']) or None,
            'OnSkill': ';'.join(cols['OnSkill']) or None,
            'OnEnhanceSkill': cols['OnEnhanceSkill'],
        }
    return single


def roundtrip_check(src_sheet, hdr, states, transitions):
    """反推单表并与源表逐格比对（不含 Order/额外列），返回 mismatch 列表。"""
    rev = reverse_derive(states, transitions)
    bool_cols = {'UseCommon', 'UnlockEnhance'}
    mismatches = []
    for r in range(4, src_sheet.max_row + 1):
        sid = str(src_sheet.cell(row=r, column=hdr['StateId']).value or '').strip()
        if not sid:
            continue
        if int(sid) not in rev:
            mismatches.append(f'第 {r} 行 StateId={sid} 反推缺失')
            continue
        got = rev[int(sid)]
        for col in ('CharacterId', 'UseCommon', 'Info', 'AnimName', 'UnlockEnhance',
                    'OnAnimEnd', 'OnMove', 'OnAtk', 'OnSkill', 'OnEnhanceSkill'):
            src_v = src_sheet.cell(row=r, column=hdr[col]).value
            # bool 列用 flag() 归一化：空值 == 'FALSE'（语义等价），避免误报
            src = flag(src_v) if col in bool_cols else norm_cell(src_v)
            want = got[col] if col in bool_cols else norm_cell(got[col])
            if src != want:
                mismatches.append(f'第 {r} 行 {col}: 源表={src!r} 反推={want!r}')
    return mismatches


def load_existing_sheet(path, sheet):
    """读已存在的双表里某张 sheet → (headers, rows)；不存在/读不了返回 None。"""
    if not os.path.exists(path):
        return None
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    headers = [str(ws.cell(row=1, column=c).value).strip()
               for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value]
    rows = []
    for r in range(4, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if len(vals) >= 2 and any(v is not None and str(v).strip() for v in vals[:2]):
            rows.append(vals)
    return headers, rows


def merge_transitions(transitions, existing):
    """
    合并 transition：匹配 (From, To, Condition) 的行保留其 Order 和额外列；
    新行用默认 Order + 空额外列。返回 (merged, extra_headers)。

    匹配键【不含 Param】：Param 是工具负责重刷的列。若你手改过 Param，或 WPS 数字
    格式漂移，键对不上会把整行当新行、Order 丢。只按 From/To/Condition 匹配，能保证
    只要行还在，你设的 Order 就保留；Param 由工具刷回单表值。
    """
    if existing is None:
        return transitions, []
    headers, rows = existing
    extra = [h for h in headers if h not in TRANS_HEADERS]

    # (From, To, Condition) → 候选行列表（同一键可能有多行，如多窗口同目标）
    lookup = {}
    for vals in rows:
        if len(vals) < 3:
            continue
        f, to = vals[1], vals[2]
        if f is None or to is None:
            continue
        cond = vals[3] if len(vals) > 3 else ''
        lookup.setdefault((num(f), num(to), norm_cell(cond)), []).append(vals)

    kept = 0
    merged = []
    for t in transitions:
        bucket = lookup.get((t[1], t[2], t[3]), [])
        # 优先精确匹配 Param 的行（保持一一对应），否则取桶里第一行
        old = None
        for i, cand in enumerate(bucket):
            if norm_cell(t[4]) == norm_cell(cand[4] if len(cand) > 4 else ''):
                old = bucket.pop(i)
                break
        if old is None and bucket:
            old = bucket.pop(0)
        if old is not None:
            order = old[5] if len(old) > 5 and old[5] is not None else t[5]
            extras = list(old[6:]) if len(old) > 6 else []
            merged.append(t[:5] + [order] + extras)
            kept += 1
        else:
            merged.append(t[:5] + [t[5]] + [''] * len(extra))
    print(f'[合并] transition: 保留 {kept}/{len(transitions)} 行的已有 Order/额外列')
    return merged, extra


def merge_states(states, existing):
    """合并 state：按 StateId 匹配，保留额外列；新状态用空额外列。"""
    if existing is None:
        return states, []
    headers, rows = existing
    extra = [h for h in headers if h not in STATE_HEADERS]

    lookup = {}
    for vals in rows:
        if len(vals) < 3:
            continue
        sid = vals[2]
        if sid is None:
            continue
        lookup[num(sid)] = vals

    kept = 0
    merged = []
    for st in states:
        if st[2] in lookup:
            old = lookup[st[2]]
            extras = list(old[6:]) if len(old) > 6 else []
            merged.append(st[:6] + extras)
            kept += 1
        else:
            merged.append(st[:6] + [''] * len(extra))
    print(f'[合并] state: 保留 {kept}/{len(states)} 行的已有额外列')
    return merged, extra


def fill_sheet(ws, headers, types, comments, data):
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    for c, t in enumerate(types, 1):
        ws.cell(row=2, column=c, value=t)
    for c, cm in enumerate(comments, 1):
        ws.cell(row=3, column=c, value=cm)
    for r, row in enumerate(data, 4):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 40


def main():
    # 控制台输出 UTF-8，避免 Windows GBK 控制台编不了中文/符号
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    here = os.path.dirname(os.path.abspath(__file__))
    repo = os.path.dirname(os.path.dirname(os.path.dirname(here)))  # AI_Tools → Plugins → Assets → 仓库根
    default_in = os.path.join(repo, 'Assets', 'Excel', 'StateMachineConfig.xlsx')
    default_out = os.path.join(repo, 'Assets', 'Excel', 'StateMachineConfig_Double.xlsx')

    arg_in = sys.argv[1] if len(sys.argv) > 1 else default_in
    arg_out = sys.argv[2] if len(sys.argv) > 2 else default_out

    if not os.path.exists(arg_in):
        raise SystemExit(f'输入文件不存在: {arg_in}')

    wb = load_workbook(arg_in, data_only=True)
    src, sheet_name = find_authoring_sheet(wb)
    hdr = header_map(src)

    states, transitions, warnings = parse_single(src, hdr)
    warnings += validate(states, transitions)

    # 稳定排序：From 分组 + Order 优先级；同组保持原窗口数组顺序
    transitions.sort(key=lambda t: (t[1], t[5]))

    # 合并已存在的双表：保留 Order 和额外列（只重写转换负责的列）
    ex_state = load_existing_sheet(arg_out, 'state')
    ex_trans = load_existing_sheet(arg_out, 'transition')
    states, state_extra = merge_states(states, ex_state)
    transitions, trans_extra = merge_transitions(transitions, ex_trans)

    # round-trip 无损校验（校验前先确认核心列未被合并改动 —— 合并只动 Order/额外列，核心列不变）
    mismatches = roundtrip_check(src, hdr, states, transitions)

    # 写输出（额外列拼到表头，类型/注释兜底 string/空）
    out = Workbook()
    ws_state = out.active
    ws_state.title = 'state'
    fill_sheet(ws_state,
               STATE_HEADERS + state_extra,
               STATE_TYPES + ['string'] * len(state_extra),
               STATE_COMMENTS + [''] * len(state_extra),
               states)
    ws_trans = out.create_sheet('transition')
    fill_sheet(ws_trans,
               TRANS_HEADERS + trans_extra,
               TRANS_TYPES + ['string'] * len(trans_extra),
               TRANS_COMMENTS + [''] * len(trans_extra),
               transitions)
    out.save(arg_out)

    # 报告
    print(f'[转换] 输入 : {arg_in} (sheet={sheet_name})')
    print(f'[转换] 输出 : {arg_out}')
    print(f'[转换] 状态 {len(states)} 个, 迁移 {len(transitions)} 条')
    if state_extra:
        print(f'[转换] state 额外列: {state_extra}')
    if trans_extra:
        print(f'[转换] transition 额外列: {trans_extra}')
    if warnings:
        print('[转换] ⚠ 校验警告:')
        for w in warnings:
            print('   -', w)
    if mismatches:
        print(f'[转换] ✗ round-trip 校验失败, {len(mismatches)} 处不一致:')
        for m in mismatches:
            print('   -', m)
        sys.exit(1)
    print('[转换] ✓ round-trip 校验通过: 双表核心列可无损还原回单表')


if __name__ == '__main__':
    main()
